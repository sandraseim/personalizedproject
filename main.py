from openpyxl import workbook, load_workbook
import xlsxwriter
from openpyxl.chart import BarChart, Reference


wb = load_workbook('Book1.xlsx')
ws = wb.active
#ws.insert_cols(2) #insert column for biology grades
print(wb.sheetnames)
sheet_name = wb['Sheet1']
sheet_name.title = 'Grades' #Sheet is renamed from Sheet1 to Grades
print(wb.sheetnames)
sheet = wb.sheetnames

tuple(ws['B2':'B6'])
cellsB = []
for rowOfCellObjects in ws['B2':'B6']:
    for cellObj in rowOfCellObjects:
        cellsB.append(cellObj.value)
av = sum(cellsB)/len(cellsB)
print(av)

tuple(ws['C2':'C6'])
cellsC = []
for rowOfCellObjectsC in ws['C2':'C6']:
    for cellObjC in rowOfCellObjectsC:
        cellsC.append(cellObjC.value)
avC = sum(cellsC)/len(cellsC)
print(avC)

tuple(ws['D2':'D6'])
cellsD = []
for rowOfCellObjectsD in ws['D2':'D6']:
    for cellObjD in rowOfCellObjectsD:
        cellsD.append(cellObjD.value)
avD = sum(cellsD)/len(cellsD)
print(avD)


tuple(ws['E2':'E6'])
cellsE = []
for rowOfCellObjectsE in ws['E2':'E6']:
    for cellObjE in rowOfCellObjectsE:
        cellsE.append(cellObjE.value)
avE = sum(cellsE)/len(cellsE)
print(avE)

#If theres no data in excel, you can add data manually in python, see below
#rows = [
#    ('Name', 'Math', 'Sport', 'English', 'Chemistry'),
#    ('Jim', 45, 65, 7, 63),
#    ('Tim', 45,	65,	7,	63),
#    ('Alex', 56, 43, 98, 76),
#    ('Zedd', 87, 65, 54, 56),
#    ('Harry', 65, 65, 41, 76)
#]
#
#for row in rows:
#    ws.append(row)

#Creating bar chart


chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Results for subject"
chart1.y_axis.title = 'result'
chart1.x_axis.title = 'student name'

data = Reference(ws, min_col=2, min_row=1, max_row=6, max_col=5)
values = Reference(ws, min_col=1, min_row=2, max_row=6)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(values)
chart1.shape = 4
ws.add_chart(chart1, "A10")





wb.save('1Book.xlsx')






# book = xlsxwriter.Workbook('1Book.xlsx')
# sheets = book.add_worksheet()
# sheets.write('B2', 'Hi')
# #row = 0
# #column = 1
# #content = ["Biology", "45", "48", "71", "67", "66"]
# ##
# #for item in content:
# #    sheet.write(row, column, item)
# #    row += 1
# #
# book.close()
# #